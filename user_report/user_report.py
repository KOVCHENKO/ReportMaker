# -------------------------------------------------------------------------------
# Name:        User_Payments
# Purpose:
#
# Author:      ITSG
#
# Created:     27.02.2019
# Copyright:   (c) ITSG 2019
# -------------------------------------------------------------------------------
import openpyxl, pymysql, os
from smtplib import SMTP_SSL
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import datetime

from settings.settings import host, user, passwd, db, port, smtp_server, mail_login, mail_passwd, receiver, cc

# Excel Settings
today = datetime.date.today().strftime('%d.%m.%Y')
excel_file = 'Oplata_polzovateley_' + today + '.xlsx'

headers = ['id', 'email', 'Контрагент', 'Тариф', 'дата оплаты', 'дата окончания', 'Группа', 'Кол-во входов',
           'тип оплаты']


def user_report():
    # Fetch Data from SQL server
    conn = pymysql.connect(host=host, user=user, passwd=passwd, db=db, port=port)
    cursor = conn.cursor()
    cursor.execute('''select contractors.title, users.id, users.email, contractors.name, rates.title, parents_payments.payment, parents_payments.finish, parents_groups.title, 
        (select count(*) from sys_log where element = 'user_login' and text = users.id) as entries
        from users
        join parents_schools on parents_schools.user_id = users.id
        join contractors on parents_schools.contractor_id = contractors.id
        join parents_payments on parents_payments.user_id = users.id
        join rates on rates.id = parents_payments.rate_id
        join parents_in_groups on parents_in_groups.parent_id = users.id
        join parents_groups on parents_groups.id = parents_in_groups.group_id
        where parents_payments.state = 'on'
        and users.state <> 'blocked'
        and CURRENT_DATE < parents_payments.finish
        order by contractors.id, parents_payments.payment''')
    data = cursor.fetchall()
    conn.close()

    # Write Data to Excel file
    wb = openpyxl.Workbook()
    contractors = {}
    for item in data:
        diff = item[6] - item[5]

        item = list(item)
        if diff.days > 10:
            item.append('полный')
        else:
            item.append('триальный')

        item = tuple(item)

        if item[0][:30] in contractors:
            contractors[item[0][:30]] += 1
        else:
            wb.create_sheet(item[0][:30])
            contractors[item[0][:30]] = 2
            for i in range(1, len(headers) + 1):
                letter = openpyxl.utils.get_column_letter(i)
                wb[item[0][:30]][letter + '1'] = headers[i - 1]
        wb[item[0][:30]]['A' + str(contractors[item[0][:30]])] = contractors[item[0][:30]] - 1
        for i in range(2, len(headers) + 1):
            letter = openpyxl.utils.get_column_letter(i)
            wb[item[0][:30]][letter + str(contractors[item[0][:30]])] = item[i]
    wb.save(excel_file)
    wb.remove(wb['Sheet'])
    wb.save(excel_file)

    # Compose attachment
    part = MIMEBase('application', "octet-stream")
    part.set_payload(open(excel_file, "rb").read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename="%s"' % os.path.basename(excel_file))

    # Compose message
    msg = MIMEMultipart()
    msg['From'] = mail_login
    msg['To'] = ', '.join(receiver)
    msg['Cc'] = ', '.join(cc)
    msg['Subject'] = excel_file
    msg.attach(part)

    # Send mail
    tosend = receiver + cc
    smtp = SMTP_SSL('smtp.gmail.com')
    smtp.connect(smtp_server)
    smtp.login(mail_login, mail_passwd)
    smtp.sendmail(mail_login, tosend, msg.as_string())
    smtp.quit()

    # Wipe file
    os.remove(excel_file)
