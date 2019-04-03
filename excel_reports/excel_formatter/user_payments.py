import datetime

import openpyxl

today = datetime.date.today().strftime('%d.%m.%Y')
excel_file = 'Oplata_polzovateley_' + today + '.xlsx'
headers = ['id', 'email', 'Контрагент', 'Тариф', 'дата оплаты', 'дата окончания', 'Группа', 'Кол-во входов',
           'тип оплаты']


def user_payments_excel_maker(data):
    wb = openpyxl.Workbook()
    contractors = {}
    for item in data:
        diff = item[6] - item[5]

        item = list(item)
        if diff.days > 10:
            item.append('полный')
        else:
            item.append('триальный')

        item = tuple(item)

        if item[0][:30] in contractors:
            contractors[item[0][:30]] += 1
        else:
            wb.create_sheet(item[0][:30])
            contractors[item[0][:30]] = 2
            for i in range(1, len(headers) + 1):
                letter = openpyxl.utils.get_column_letter(i)
                wb[item[0][:30]][letter + '1'] = headers[i - 1]
        wb[item[0][:30]]['A' + str(contractors[item[0][:30]])] = contractors[item[0][:30]] - 1
        for i in range(2, len(headers) + 1):
            letter = openpyxl.utils.get_column_letter(i)
            wb[item[0][:30]][letter + str(contractors[item[0][:30]])] = item[i]
    wb.save(excel_file)
    wb.remove(wb['Sheet'])
    wb.save(excel_file)
